import os
import time
import nibabel as nib
import numpy as np
import scipy.spatial.distance as distance
import openpyxl
import matplotlib.pyplot as plt

import img_processing_functions as img_func
import utils.img_utils as img_utils

import warnings
warnings.filterwarnings("ignore")


def load_subject_data(folder_path, mask=False, display=False):
    if not mask:
        image_file = os.path.join(
            folder_path, os.path.basename(folder_path) + "_ana.nii.gz")
    else:
        image_file = os.path.join(
            folder_path, os.path.basename(folder_path) + "_ana_brainmask.nii.gz")
    image_data = nib.load(image_file)
    raw_img = image_data.get_fdata()
    pixel_dims = (raw_img.shape[0], raw_img.shape[1], raw_img.shape[2])
    raw_images = np.zeros(pixel_dims, dtype=image_data.get_data_dtype())

    for index_1 in range(pixel_dims[0]):
        for index_2 in range(pixel_dims[1]):
            for index_3 in range(pixel_dims[2]):
                raw_images[index_1, index_2,
                           index_3] = raw_img[index_1, index_2, index_3][0]
    if not mask:
        raw_images = np.rot90(raw_images, axes=(0, 1))
        raw_images = np.rot90(raw_images, axes=(1, 2))
        raw_images = np.rot90(raw_images, k=2, axes=(2, 0))
    else:
        raw_images = np.rot90(raw_images, axes=(0, 2))
        raw_images = np.rot90(raw_images, k=2, axes=(1, 0))

    if display:
        img_utils.plot_stack(os.path.basename(folder_path), raw_images)
    return raw_images


def calc_jaccard(mask1, mask2):
    mask1 = np.asarray(mask1).astype(np.bool)
    mask2 = np.asarray(mask2).astype(np.bool)
    intersection = np.logical_and(mask1, mask2)
    union = np.logical_or(mask1, mask2)

    jaccard = intersection.sum() / float(union.sum())
    return jaccard


def calc_metrics(subject_id, masks, ground_truth_mask):
    mask = masks["consensus"]
    n_pixels_brain = np.sum(ground_truth_mask)
    tp = 0
    fn = 0
    fp = 0
    for ind_1 in range(0, ground_truth_mask.shape[0]):
        for ind_2 in range(0, ground_truth_mask.shape[1]):
            for ind_3 in range(0, ground_truth_mask.shape[2]):
                if ground_truth_mask[ind_1, ind_2, ind_3] == mask[ind_1, ind_2, ind_3]:
                    if ground_truth_mask[ind_1, ind_2, ind_3]:
                        tp += 1
                else:
                    if mask[ind_1, ind_2, ind_3]:
                        fp += 1
                    else:
                        fn += 1

    tpr = tp/n_pixels_brain
    fpr = fp/n_pixels_brain
    fnr = fn/n_pixels_brain
    jaccard = calc_jaccard(ground_truth_mask, mask)
    # jaccard = distance.jaccard(ground_truth_mask, mask)
    return [subject_id, tpr, fpr, fnr, jaccard]


def save_metrics_excel(metrics, xlsx_path):
    wb = openpyxl.Workbook()
    ws = wb.active

    for row in metrics:
        ws.append(row)

    wb.save(xlsx_path)


def load_metrics_excel(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    metrics = []
    for row in range(1, ws.max_row + 1):
        row_vals = []
        for col in range(2, ws.max_column + 1):
            row_vals.append(float(ws.cell(row, col).value))
        metrics.append(row_vals)
    return metrics


def create_boxplots(xlsx_path):
    metrics = load_metrics_excel(xlsx_path)
    metrics = np.array(metrics)

    # Create plot
    fig = plt.figure(figsize=(10, 7))
    ax = fig.add_subplot(111)
    bp = ax.boxplot(metrics, patch_artist=True)
    ax.set_xticklabels(['TP', 'FPR', 'FNR', 'Jaccard'])

    box_color = "#85C1E9"
    for patch in bp['boxes']:
        patch.set_facecolor(box_color)

    for median in bp['medians']:
        median.set(color='red', linewidth=2)

    plt.show()


if __name__ == "__main__":
    # For each subject in eval_database
    #   Load brain volume
    #   Load available mask
    #   Compute mask with implemented mcstrip algorithm
    #   Compute selected metrics and save
    # Create Excel file with obtained metric values
    # Generate comparation figure with 3 subjects -> boxplot

    # READ DATABASE
    eval_database_path = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), "database_eval")
    subject_path_list = [os.path.join(eval_database_path, folder) for folder in os.listdir(
        eval_database_path) if os.path.isdir(os.path.join(eval_database_path, folder))]
    subject_path_list.sort()

    xlsx_path = "evaluation/eval_results.xlsx"
    metrics = []
    subject_ids = []
    subject_img = {}
    subject_brainmask_ground_truth = {}
    for subject_path in subject_path_list:
        subject_id = os.path.basename(subject_path)
        # Read brain volume
        subject_img[subject_id] = {}
        subject_img[subject_id]['t1'] = load_subject_data(subject_path)
        # Read brain mask
        subject_brainmask_ground_truth[subject_id] = {}
        subject_brainmask_ground_truth[subject_id]['t1'] = load_subject_data(
            subject_path, mask=True)
        subject_ids.append(subject_id)

    subject_brainmasks_mcstrip = {}
    subject_masked_slices = {}
    for s in subject_ids:
        print(subject_id)
        # Compute brain mask with McStrip algorithm
        subject_brainmasks_mcstrip[s],  subject_masked_slices = img_func.image_skull_strip(
            subject_img[s]['t1'], False)
        img_utils.plot_stack(subject_id, subject_masked_slices)
        metrics.append(calc_metrics(
            s, subject_brainmasks_mcstrip[s], subject_brainmask_ground_truth[s]['t1']))

    save_metrics_excel(metrics, xlsx_path)
    # print(metrics)
    create_boxplots(xlsx_path)
